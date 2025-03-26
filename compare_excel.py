import openpyxl
from openpyxl.styles import Font, PatternFill

def compare_excel_files_by_keys(file1, file2, output_file):
    # Load both workbooks
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # Create a new workbook for the result
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = "Comparison Results"

    # Compare sheets
    for sheet_name in wb1.sheetnames:
        if sheet_name not in wb2.sheetnames:
            print(f"Sheet '{sheet_name}' not found in the second file.")
            continue

        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]

        # Create dictionaries to store rows by key (first two columns)
        data1 = {}
        data2 = {}

        for row in ws1.iter_rows(min_row=2, values_only=True):
            key = (row[0], row[1])  # First two columns as key
            data1[key] = row

        for row in ws2.iter_rows(min_row=2, values_only=True):
            key = (row[0], row[1])  # First two columns as key
            data2[key] = row

        # Write headers to the output sheet
        headers = [cell.value for cell in ws1[1]]
        headers.extend(["filename", "same"])
        if ws_output.max_row == 1:  # Write headers only once
            ws_output.append(headers)

        # Compare rows and write results
        for key in set(data1.keys()).union(data2.keys()):
            row1 = data1.get(key)
            row2 = data2.get(key)
            same_value = 1

            # Write file1 row
            if row1:
                output_row = list(row1) + ["file1"]
                if row2:
                    # Compare values for the same key
                    for col_num, (val1, val2) in enumerate(zip(row1, row2), start=3):
                        if val1 != val2:
                            same_value = 0
                            ws_output.cell(row=ws_output.max_row + 1, column=col_num).font = Font(bold=True, color="FF0000")
                else:
                    same_value = 0  # No match in file2
                output_row.append(same_value)
                ws_output.append(output_row)

            # Write file2 row
            if row2:
                output_row = list(row2) + ["file2"]
                if row1:
                    # Highlight differences for the same key
                    for col_num, (val1, val2) in enumerate(zip(row1, row2), start=3):
                        if val1 != val2:
                            same_value = 0
                            ws_output.cell(row=ws_output.max_row + 1, column=col_num).font = Font(bold=True, color="FF0000")
                else:
                    same_value = 0  # No match in file1
                output_row.append(same_value)
                ws_output.append(output_row)

    # Save the result to the output file
    wb_output.save(output_file)
    print(f"Comparison completed. Results saved to '{output_file}'.")

# Usage
file1 = "GBI083_OPTI.xlsx"
file2 = "GBI083_RAW.xlsx"
output_file = "comparison_result.xlsx"

compare_excel_files_by_keys(file1, file2, output_file)